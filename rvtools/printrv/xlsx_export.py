"""Module for exporting data to XLSX format with multiple sheets"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class XlsxExporter:
    """Export multiple sheets to XLSX workbook"""

    def __init__(self, filename, directory):
        """Initialize XLSX exporter"""
        self.filename = filename
        self.directory = directory
        self.full_path = f"{directory}/{filename}"
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

    def add_sheet(self, sheet_name, data):
        """Add a sheet with data to the workbook"""
        if not data:
            return

        sheet = self.workbook.create_sheet(sheet_name)

        # Get headers from first row
        headers = list(data[0].keys())

        # Write headers with styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = row_data.get(header, "")
                
                # Convert value to Excel-compatible format
                if value is None:
                    cell.value = ""
                elif isinstance(value, (str, int, float, bool, datetime)):
                    cell.value = value
                else:
                    # Convert other types to string
                    try:
                        cell.value = str(value)
                    except Exception:
                        cell.value = "[Error converting value]"
                
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=False
                )

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = adjusted_width

        # Freeze header row
        sheet.freeze_panes = "A2"

    def save(self):
        """Save the workbook"""
        print(f"## Creating {self.full_path} file.")
        self.workbook.save(self.full_path)
        print(f"✓ XLSX file saved: {self.full_path}")
